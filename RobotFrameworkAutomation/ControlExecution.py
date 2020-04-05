import openpyxl
excel = openpyxl.load_workbook('C:/Users/Nitin Tyagi/Desktop/task/Data2.xlsx')
sh = excel['Sheet1']
def  check_selected_country(countyrName):
    z="hello"
    if(countyrName=="Apr"):
        z="hi"
    return str(z)


def read_excel_rows_count():
    r= sh.max_row
    r=r+1
    return  r

def read_excel_column_count():
    c= sh.max_column
    return  c

def read_excel_cell_data(r,c):
    cl= sh.cell(int(r),int(c))
    return cl.value

def write_excel_cell_data(r,c,data):
    cl= sh.cell(int(r),int(c))
    cl.value = data